import os
import pandas as pd
import openpyxl
from datetime import datetime
from selenium import webdriver
from time import sleep
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.edge.service import Service
import tkinter as tk
import send2trash
from Oddo_Extract_Sales import Extract_sales
from Oddo_Extract_Sales import definir_col
from Oddo_Extract_Sales import cedula_branco

from tkinter import messagebox

today = datetime.today()
year = int(today.date().year)
month = int(today.date().month)
day = int(today.date().day)

if month <= 9:
    month_string = f"0{month}"
else:
    month_string = f"{month}"

if day <= 9:
    day_string = f"0{day}"
else:
    day_string = f"{day}"

data_texto = f"{day_string}/{month_string}/{year}"
Link_odoo = ""

print("Antes de iniciar o script tenha em vista algumas necessidades:\n")

print("1. É necessário que você tenha o sharepoint OTC P&SP Brazil - NF_Remessa_Automation sincronizado em seu computador:\nLink: https://workspaces.bsnconnect.com/teams/OTC_Brasil/Shared%20Documents/NF_Remessa_Automation")
print("\n")

print("2. Este arquivo executável assim como o msedgedriver.exe devem estar no DESKTOP do seu computador;")
print("\n")

print("3. Sempre antes de iniciar o código verifique a pasta Automatizacao -> _Data_Control_-> Presenca_Carga_Email. Nela devem estar presentes as planilhas de presença de carga;")
print("OBS: Garanta que caso você rode o código mais vezes para as mesmas planilhas, que elas estejam sempre presentes na pasta Presenca_Carga_Email;")
print("\n")

print("4. Também verifique se o arquivo id_customer_CC.xlsx está atualizado. Uma vez que ele não esteja, não serão encontrados os Cost Centers e, esses casos não estarão presentes na Carga massiva;")
print("\n")

print("5. No momento de importar a Carga massiva, apenas selecione o arquivo Carga_massiva.xlsx, que está na pasta _Data_Controls -> Carga_massiva, e clique Open;")
print("OBS: Você terá 1 minuto para escolher o arquivo e após escolhido, não clique em nada, pois o código fará a importação no sistema;")
print("\n")

print("6. Após o encerramento do código em _Data_Control_ -> Database_SalesOrder, você encontrará um arquivo excel com todas as informações e uma pasta NFs_PDFs que contém as NFs;")
print("OBS: Não deixe de considerar a pasta Carga_massiva, pois nela há uma planilha que destaca os casos que não foram encontrados Cost Centers e, portanto, não foram importados para o Odoo;")
print("\n")

print("7. Por fim, para extração das NFs garanta que o SAP esteja totalmente fechado ou aberto na tela principal")

print("\n")



user_id = input("Digite seu main number:")
Odoo_user = input("Digite seu e-mail de login no Odoo:")
Odoo_password = input("Digite sua senha do Odoo:")
subida_arquivo = input('Você deseja subir a carga massiva (1) ou apenas verificar os casos de faturamento (2)?')

#inicializando arquivo que será editado


import openpyxl

# Caminho para o arquivo Excel
file_path = f"C:\\Users\\{user_id}\\Company\\NF_Remessa_Automation\\Automatizacao\\_Data_Control_\\carga_massiva\\Carga_massiva.xlsx"
cost_center_path = f"C:\\Users\\{user_id}\\Company\\NF_Remessa_Automation\\Automatizacao\\_Data_Control_\\id_customer_CC.xlsx"
file_pathwithoutCC = f"C:\\Users\\{user_id}\\Company\\NF_Remessa_Automation\\Automatizacao\\_Data_Control_\\carga_massiva\\Carga_massiva_sem_cc.xlsx"
#tabela cost center: 
cost_center = pd.read_excel(cost_center_path, sheet_name = "Sheet1")

#Obtendo informações enviadas por fornecedor
# Caminho da pasta
directory = f"C:\\Users\\{user_id}\\Company\\NF_Remessa_Automation\\Automatizacao\\_Data_Control_\\Presenca_Carga_Email"

# Carrega a planilha
workbook = openpyxl.load_workbook(file_path)
workbook2 = openpyxl.load_workbook(file_pathwithoutCC)
# Seleciona a planilha ativa
sheet = workbook.active
sheet2 = workbook2.active

# Exclui todas as linhas na planilha
sheet.delete_rows(2, sheet.max_row)
sheet2.delete_rows(2, sheet2.max_row)

# Inicializa a contagem
excel_count = 0
excel_files = []

try:

    # Percorre os arquivos na pasta
    for filename in os.listdir(directory):
        # Verifica se o arquivo é uma planilha Excel
        if filename.endswith('.xlsx'):
            excel_count += 1
            excel_files.append(filename)

        # Exibe a lista de arquivos Excel

except FileNotFoundError:
    print(f"A pasta '{directory}' não existe.")

# Exibe a contagem de arquivos Excel
print(f"Há {excel_count} planilhas na pasta.")
print(f"{excel_files}")

for i in range(0,excel_count,1):

    if i == 0:

        definir_col(directory,excel_files[i], sheet, sheet2, cost_center)
        #cedula_branco(linhas, sheet3, i, excel_files, workbook3,directory)
 
      
            
    else:
        definir_col(directory,excel_files[i], sheet, sheet2, cost_center)
        #cedula_branco(linhas, sheet3, i, excel_files, workbook3,directory)

#Salva a planilha
workbook.save(file_path)
workbook2.save(file_pathwithoutCC)


df_carga_massiva = pd.read_excel(file_path)
df_carga_massiva_CC = pd.read_excel(file_pathwithoutCC)


df_unique = df_carga_massiva.drop_duplicates(subset=["Observações"], keep="first")
df_unique_CC = df_carga_massiva_CC.drop_duplicates(subset=["Observações"], keep="first")



linhas_total = len(df_unique)
linhas_total_CC = len(df_unique_CC)
print(f"qtde total de linhas - Carga Massiva: {linhas_total}")
print(f"qtde total de linhas - Carga Massiva sem CC: {linhas_total_CC}")


for i in range(len(df_unique)):
    df_unique.loc[df_unique.index[i], "Id Solicitação"] = i + 1
    df_unique.loc[df_unique.index[i], "Linha do Item"] = i + 1


df_unique.to_excel(file_path, index=False)


   



#interface com Odoo:

# Instantiate the webdriver with the executable location of MS Edge web driver
driver_path = f"C:\\Users\\{user_id}\\Company\\Desktop\\msedgedriver.exe"
service = Service(executable_path=driver_path)

browser = webdriver.Edge(service=service)
# Simply just open a new Edge browser and go to lambdatest.com
browser.maximize_window()
browser.get(Link_odoo)

user = browser.find_element(By.XPATH, '//*[@id="login"]')
sleep(2)
user.send_keys(Odoo_user)
sleep(1)
        
senha = browser.find_element(By.XPATH, '//*[@id="password"]')
sleep(2)
senha.send_keys(Odoo_password)
sleep(1)

Login_button = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrapwrap"]/main/div/div/div/form/div[3]/button')))
Login_button.click()
sleep(8)

home_button = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '/html/body/header/nav/div[1]/button')))
home_button.click()
sleep(2)

carga_massiva = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '/html/body/header/nav/div[1]/div/span[2]/a')))
carga_massiva.click()
sleep(3)

if subida_arquivo == "1":
    Favoritos = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[1]/div[2]/div[2]/div[1]/div[3]/button')))
    Favoritos.click()
    sleep(3)

    import_registro = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[1]/div[2]/div[2]/div[1]/div[3]/div/span')))
    import_registro.click()
    sleep(3)

    carregar_arquivo = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[1]/div[2]/div[1]/div/button[4]')))
    carregar_arquivo.click()

    sleep(60)

    teste = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[1]/div[2]/div[1]/div/button[3]')))
    teste.click()

    elemento = WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/form/div[1]/div[2]/div[3]/div/span")))
    texto = elemento.text

    root = tk.Tk()
    root.withdraw()  # Oculta a janela principal

    if texto.strip() == "Tudo parece válido.":
    
        import_button = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[1]/div[2]/div[1]/div/button[1]')))
    
        import_button.click()       
    else:
        browser.close()

    sleep(4)
else:
    pass

print("Verificando casos da Carga Massiva - Atualizado ou Cancelado...")
j = 1

while True:
    sleep(5)
    browser.refresh()
    sleep(3)

    status_cargamassiva = []  # cria uma lista vazia

    for j in range(1, linhas_total + 1):  # percorre todas as linhas
        status_element = WebDriverWait(browser, 30).until(
        EC.visibility_of_element_located((By.XPATH, f'/html/body/div[1]/div/div[2]/div/div[1]/table/tbody/tr[{j}]/td[12]'))
    )
        status_text = status_element.text.strip()
        status_cargamassiva.append(status_text)


    if all(status_text1 in ("Cancelado", "Atualizado", "Rejeitado") for status_text1 in status_cargamassiva) == True:
        break

    else:
        sleep(50)
        pass



print("Casos verificados!")

status_cargamassiva.reverse()
print(status_cargamassiva)



print("Casos verificados!")

status_cargamassiva.reverse()

tabela = pd.read_excel(file_path)

# Depois de coletar os status
print("Itens coletados:", len(status_cargamassiva))

# Atribuir ao DataFrame
tabela['Status - Carga Massiva'] = status_cargamassiva

sleep(5)

home_button = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '/html/body/header/nav/div[1]/button')))
home_button.click()
sleep(2)

solicitacoes = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, '/html/body/header/nav/div[1]/div/span[1]/a')))
solicitacoes.click()
sleep(3)

print("Verificando faturamento...")
i = 1
contador = 1
while True:
    sleep(5)
    browser.refresh()
    sleep(3)

    status_total = []
    for i in range(1,linhas_total + 1):
        if tabela.iloc[i-1]["Status - Carga Massiva"] == "Cancelado" or tabela.iloc[i-1]["Status - Carga Massiva"] == "Rejeitado":
            status_text = 'Cancelado'
            status_total.append(status_text)
        else:
            status =  WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.XPATH, f'/html/body/div[1]/div/div[2]/div/div[1]/table/tbody/tr[{i}]/td[6]')))
            status_text = status.text.strip()
            status_total.append(status_text)
            print(f" status - solicitacao: {status_text}")
            sleep(2)

            if status_text.strip() == "Faturamento Pendente" and i == 1:
                
                browser.execute_script("arguments[0].scrollIntoView({block: 'center'});", status)
                status.click()
                sleep(1)
                faturar = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.NAME, "action_invoice_order")))
                faturar.click()
                sleep(1)
                #ok_button = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH,f'//*[@id="modal_{first+(61*(contador-1))}"]/div/div/footer/button[1]')))#modal_220 > div > div > footer > button.btn.btn-primary > span
                ok_button = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.CSS_SELECTOR,f'div.modal.show footer button.btn.btn-primary')))#modal_220 > div > div > footer > button.btn.btn-primary > span
                ok_button.click()
                contador += 1
            
            elif status_text.strip() == "Faturamento Pendente" and i > 1:
                
                browser.execute_script("arguments[0].scrollIntoView({block: 'center'});", status)
                status.click()
                sleep(1)
                faturar = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.NAME, "action_invoice_order")))
                faturar.click()                       
                sleep(1)
                #ok_button = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="modal_{first+(61*(contador-1))}"]/div/div/footer/button[1]')))
                #ok_button.click()
                ok_button = WebDriverWait(browser, 40).until(EC.element_to_be_clickable((By.CSS_SELECTOR, f'div.modal.show footer button.btn.btn-primary')))  # modal_220 > div > div > footer > button.btn.btn-primary > span
                ok_button.click()
                contador += 1
            elif status_text.strip() == "Aguardando criação da Nota":
                pass
            else:
                pass

    if any(status_text in ("Faturamento Pendente", "Aguardando criação de Ordem") for status_text in status_total) == False:
        break

    else:
        sleep(50)
        pass

Solicitacao_ID = []
billing = []
SalesOrder = []

for j in range(1, linhas_total + 1):  # percorre todas as linhas
    billing_element = WebDriverWait(browser, 30).until(EC.visibility_of_element_located((By.XPATH, f'/html/body/div[1]/div/div[2]/div/div[1]/table/tbody/tr[{j}]/td[9]')))
    billing_text = billing_element.text.strip()
    billing.append(billing_text)

    solicitacao_element = WebDriverWait(browser, 30).until(EC.visibility_of_element_located((By.XPATH, f'/html/body/div[1]/div/div[2]/div/div[1]/table/tbody/tr[{j}]/td[2]')))
    solcitacao_text = solicitacao_element.text.strip()
    Solicitacao_ID.append(solcitacao_text)

    salesOrder_element = WebDriverWait(browser, 30).until(EC.visibility_of_element_located((By.XPATH, f'/html/body/div[1]/div/div[2]/div/div[1]/table/tbody/tr[{j}]/td[8]')))
    salesOrder_text = salesOrder_element.text.strip()
    SalesOrder.append(salesOrder_text)

sleep(5)


print("Faturamento completado")

tabela["Solicitacao ID"] = Solicitacao_ID
tabela["Sales Order"] = SalesOrder
tabela["Billing"] = billing


Extract_sales(browser, linhas_total, user_id, data_texto, tabela)

#Excluir planilhas
'''
try:
    # Percorre os arquivos na pasta
    for filename in os.listdir(directory):
        # Verifica se o arquivo é uma planilha Excel
        if filename.endswith('.xlsx'): #or filename.endswith('.xls')
            send2trash.send2trash(f'{directory}\\{filename}')
    print("Presenças de cargas utilizadas excluídas com Sucesso!")
except FileNotFoundError:
    print(f"A pasta '{directory}' não existe.")
'''
root = tk.Tk()
root.withdraw()  # Oculta a janela principal

sleep(5)
browser.close()
sleep(2)
messagebox.showinfo("Sucesso", "A automatização foi concluída com sucesso!")

root.destroy()